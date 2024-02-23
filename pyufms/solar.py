""" Helper script for downloading data from sems portal and ploting graphs """
import json
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import requests
import xlrd
from pyufms.config import ARGS, INVERTERS_INFO
from pyufms.inverters import Inverter, UC
from pyufms.plot import plot_inverter_kwh_for_day, plot_inverter_kwh_for_month

API_URL = "https://www.semsportal.com/api/"

headers = {"Token": "{'version': 'v2.1.0', 'client': 'ios', 'language': 'en'}"}


def login() -> None:
    payload = {
        "account": ARGS.get("gw_account"),
        "pwd": ARGS.get("gw_password"),
    }
    url = API_URL + "v1/Common/CrossLogin"
    response = requests.post(url, headers=headers, json=payload)
    Token = response.json().get("data")
    headers["Token"] = json.dumps(Token)


def get_status(inverter: Inverter) -> dict:
    inverter_info = INVERTERS_INFO.get(inverter)
    if inverter_info is None:
        raise Exception("Error getting inverter's info: " + inverter.name)
    station_id = inverter_info.get("station_id")
    if station_id is None:
        raise Exception("Error getting inverter's station_id: " + inverter.name)
    sn = inverter_info.get("sn")
    if sn is None:
        raise Exception("Error getting inverter's sn: " + inverter.name)

    url = API_URL + "v3/PowerStation/GetInverterPoint"
    payload = {"powerStationId": station_id, "sn": sn}
    response = requests.post(url, headers=headers, json=payload)
    data = response.json().get("data")[0]
    return data


def verify_inv_names() -> None:
    login()
    for inverter in Inverter:
        inv_info = INVERTERS_INFO.get(inverter)
        if inv_info:
            fake_name = inv_info.get("fake_name")
            sn = inv_info.get("sn")
            if sn is not None:
                name = get_status(inverter).get("name")
                print(str(name) + "  <--->  " + str(fake_name))


def get_excel_payload(start: datetime, end: datetime, station_id: str, sn: str) -> dict:
    return {
        "tm_content": {
            "qry_time_start": str(start),
            "qry_time_end": str(end),
            "pws_historys": [
                {
                    "id": station_id,
                    "inverters": [{"sn": sn}],
                }
            ],
            "targets": [
                {"target_key": "Vpv1", "target_index": 1},
                {"target_key": "Vpv2", "target_index": 2},
                {"target_key": "Vpv3", "target_index": 3},
                {"target_key": "Vpv4", "target_index": 4},
                {"target_key": "Ipv1", "target_index": 5},
                {"target_key": "Ipv2", "target_index": 6},
                {"target_key": "Ipv3", "target_index": 7},
                {"target_key": "Ipv4", "target_index": 8},
                {"target_key": "Vac1", "target_index": 9},
                {"target_key": "Vac2", "target_index": 10},
                {"target_key": "Vac3", "target_index": 11},
                {"target_key": "Iac1", "target_index": 12},
                {"target_key": "Iac2", "target_index": 13},
                {"target_key": "Iac3", "target_index": 14},
                {"target_key": "Fac1", "target_index": 15},
                {"target_key": "Fac2", "target_index": 16},
                {"target_key": "Fac3", "target_index": 17},
                {"target_key": "Pac", "target_index": 18},
                {"target_key": "WorkMode", "target_index": 19},
                {"target_key": "Tempperature", "target_index": 20},
                {"target_key": "ETotal", "target_index": 22},
                {"target_key": "HTotal", "target_index": 23},
                {"target_key": "Istr1", "target_index": 24},
                {"target_key": "Istr2", "target_index": 25},
                {"target_key": "Istr3", "target_index": 26},
                {"target_key": "Istr4", "target_index": 27},
                {"target_key": "Istr5", "target_index": 28},
                {"target_key": "Istr6", "target_index": 29},
                {"target_key": "Istr7", "target_index": 30},
                {"target_key": "Istr8", "target_index": 31},
                {"target_key": "Istr10", "target_index": 33},
                {"target_key": "Istr11", "target_index": 34},
                {"target_key": "Reserved5", "target_index": 36},
                {"target_key": "PF", "target_index": 407},
                {"target_key": "ReactivePower", "target_index": 408},
                {"target_key": "LeakageCurrent", "target_index": 411},
                {"target_key": "ISOLimit", "target_index": 412},
            ],
        }
    }


def get_excel_qry_key(start: datetime, end: datetime, inverter: Inverter) -> str:
    inverter_info = INVERTERS_INFO.get(inverter)
    if inverter_info is None:
        raise Exception("Error getting inverter's info: " + inverter.name)

    station_id = inverter_info.get("station_id")
    if station_id is None:
        raise Exception("Error getting inverter's station_id: " + inverter.name)

    sn = inverter_info.get("sn")
    if sn is None:
        raise Exception("Error getting inverter's sn: " + inverter.name)

    url = API_URL + "HistoryData/ExportExcelStationHistoryData"
    payload = get_excel_payload(start, end, station_id, sn)
    response = requests.post(url, headers=headers, json=payload)
    data = response.json().get("data")
    qry_key = data.get("qry_key")
    return qry_key


def get_excel_file_path(start: datetime, end: datetime, inverter: Inverter) -> str:
    qry_key = get_excel_qry_key(start, end, inverter)
    url = API_URL + "HistoryData/GetStationHistoryDataFilePath"
    payload = {"file_id": qry_key}
    response = requests.post(url, headers=headers, json=payload)
    data = response.json().get("data")
    file_path = data.get("file_path")
    return file_path


def clean_excel(xls_path: str, inverter: Inverter) -> Path:
    xls_book = xlrd.open_workbook(xls_path)
    xls_sheet = xls_book.sheet_by_index(0)

    xlsx_book = openpyxl.Workbook()
    xlsx_sheet = xlsx_book.active
    xlsx_sheet.title = inverter.name

    # TABLE HEADER
    row = xls_sheet.row(2)
    cells = len(row)
    for i in range(cells):
        cell = str(row[i]).split("'")[1]
        xlsx_sheet.cell(row=1, column=i + 1).value = cell

    # CELLS
    for xls_row in range(3, xls_sheet.nrows):
        row = xls_sheet.row(xls_row)

        time_str = str(row[0]).split("'")[1]
        time_obj = datetime.strptime(time_str, "%d/%m/%y %H:%M:%S")
        xlsx_sheet.cell(row=xls_row - 1, column=1).value = time_obj

        cells = len(row)
        for i in range(1, cells):
            cell = str(row[i]).split("'")[1]
            cell_text = cell
            try:
                cell = float(cell)
            except ValueError:
                cell = cell_text

            xlsx_sheet.cell(row=xls_row - 1, column=i + 1).value = cell

    path_to_save = Path(xls_path.split(".")[0] + ".xlsx")
    xlsx_book.save(path_to_save)
    Path.unlink(Path(xls_path))
    return path_to_save


def get_excel(start: datetime, end: datetime, inverter: Inverter) -> Path:
    file_path = get_excel_file_path(start, end, inverter)
    response = requests.get(file_path)
    response.raise_for_status()
    path_to_save = Path(start.strftime("%d_%m_%Y-") + inverter.name + ".xls")
    path_to_save.write_bytes(response.content)
    try:
        return clean_excel(path_to_save.name, inverter)
    except Exception:
        raise Exception(
            "Error cleaning XLS, probably empty file:  " + path_to_save.name
        )


def publish_inverter_data_for_day(day: datetime, inverter: Inverter) -> None:
    end = datetime(day.year, day.month, day.day, 22, 0, 0)
    try:
        xlsx_path = get_excel(day, end, inverter)
    except Exception as e:
        raise Exception("Error publishing " + inverter.name + "\n" + str(e)) from e

    image_path = plot_inverter_kwh_for_day(xlsx_path.name)

    # create dir for storing data
    uc = inverter.name.split("_")[0].lower()
    dir = (
        "data/{}/".format(uc)
        + str(day.year)
        + "/"
        + str(day.month)
        + "/"
        + str(day.day)
        + "/"
        + inverter.name
        + "/"
    )
    Path(dir).mkdir(parents=True, exist_ok=True)

    # move xlsx to corresponding dir
    xlsx_path_new = Path(dir + xlsx_path.name)
    xlsx_path.rename(xlsx_path_new)

    # move image (plot) to corresponding dir
    image_path_new = Path(dir + image_path.name)
    image_path.rename(image_path_new)


def publish_uc_data_for_day(uc: UC, day: datetime, skip: tuple) -> None:
    publish_count = 0
    publish_max = 0
    print("Start publishing data for {} - ".format(uc.name) + day.strftime("%d/%m/%Y"))
    for inverter in Inverter:
        if inverter in skip:
            continue

        if uc.name != inverter.name.split("_")[0]:
            continue

        publish_max += 1

        inv_info = INVERTERS_INFO.get(inverter)
        if inv_info:
            sn = inv_info.get("sn")
            if sn is not None:
                for attempt in range(5):
                    try:
                        print("Publishing data for: " + inverter.name)
                        publish_inverter_data_for_day(day, inverter)
                        update_inverter_readme_for_day(day, inverter)
                        publish_count += 1
                    except Exception as e:
                        print(e)
                        print("Retrying...")
                    else:
                        break
    print(
        "published {}/{} inverter data for {}".format(
            publish_count, publish_max, uc.name
        )
    )


def update_inverter_readme_for_day(day: datetime, inverter: Inverter) -> None:
    uc = inverter.name.split("_")[0].lower()
    dir = (
        "data/{}/".format(uc)
        + str(day.year)
        + "/"
        + str(day.month)
        + "/"
        + str(day.day)
        + "/"
        + inverter.name
        + "/"
    )
    xlsx_name = day.strftime("%d_%m_%Y-") + inverter.name + ".xlsx"
    xlsx_path = Path(dir + xlsx_name)

    book = openpyxl.load_workbook(xlsx_path)
    sheet = book.active

    # get kWh for a day
    kWh_column = 0
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == "Total Generation(kWh)":
            kWh_column = col
            break
    kWh_0 = sheet.cell(row=2, column=kWh_column).value
    kWh_1 = sheet.cell(row=sheet.max_row, column=kWh_column).value
    kWh = kWh_1 - kWh_0

    # get power
    power = 0
    data = get_status(inverter)
    data_dict = data.get("dict", {})
    data_list = data_dict.get("left")
    for i in data_list:
        if i.get("key") == "DeviceParameter_capacity":
            power = i.get("value")
            break

    # image path
    image_name = xlsx_name.split(".")[0] + ".png"

    # markdown
    markdown_table_1 = "| Inversor | Dia | Potência | kWh    |\n"
    markdown_table_2 = "| -------- | --- | -------- | ------ |\n"
    markdown_table_3 = "| {}       | {}  | {}       | {:.2f} |".format(
        inverter.name, day.date(), power, kWh
    )
    markdown_plot = "![My Image]({})\n".format(image_name)
    markdown_str = (
        markdown_plot + markdown_table_1 + markdown_table_2 + markdown_table_3
    )
    markdown_path = Path(dir + "README.md")
    markdown_path.write_text(markdown_str)


def update_uc_readme_for_month(month: datetime, uc_kWh_month_data: dict) -> None:
    uc = uc_kWh_month_data.get("uc")
    uc = str(uc).lower()
    dir = "data/{}/".format(uc) + str(month.year) + "/" + str(month.month) + "/"
    Path(dir + "plots/").mkdir(parents=True, exist_ok=True)

    markdown_path = Path(dir + "README.md")
    summary = "# Resumo\n"
    summary += "| Inversor | kWh    |\n| -------- | ------ |\n"
    plots = "# Geração Mensal por Inversor\n"

    inverters = uc_kWh_month_data.get("inverters", {})

    for inverter in inverters.keys():
        inv_kWh_month_data = inverters.get(inverter, {})
        inv_kWh_month = inv_kWh_month_data.get("total", 0)

        ## summary
        summary_line = "| {}       | {:.2f} |\n".format(inverter.name, inv_kWh_month)
        summary += summary_line

        ## plots
        image_path_relative = "plots/" + inverter.name + ".png"
        image_path = dir + image_path_relative
        plot_inverter_kwh_for_month(image_path, inv_kWh_month_data)
        plot_line = "## {}\n![My Image]({})\n".format(
            inverter.name, image_path_relative
        )
        plots += plot_line

    kWh_total = uc_kWh_month_data.get("total", 0)
    summary_line = "| {}       | {:.2f} |\n".format("kWh_total", kWh_total)
    summary += summary_line

    markdown_path.write_text(summary + plots)


def get_uc_kwh_for_month(uc: UC, month: datetime, skip: tuple) -> dict:
    uc_kWh_month_data = {}
    inverters = {}
    kWh_total = 0

    for inverter in Inverter:
        if inverter in skip:
            continue

        if uc.name != inverter.name.split("_")[0]:
            continue

        inv_kWh_month_data = get_inverter_kwh_for_month(month, inverter)
        inv_kWh_month = inv_kWh_month_data.get("total", 0)
        inverters.update({inverter: inv_kWh_month_data})
        kWh_total += inv_kWh_month

    uc_kWh_month_data.update({"uc": uc.name})
    uc_kWh_month_data.update({"inverters": inverters})
    uc_kWh_month_data.update({"total": kWh_total})

    return uc_kWh_month_data


def get_inverter_kwh_for_month(month: datetime, inverter: Inverter) -> dict:
    inv_kWh_month_data = {"total": 0, "days": {}}
    days = {}
    kWh_total = 0
    kWh_month_0 = 0
    kWh_month_1 = 0

    day = datetime(month.year, month.month, 1)
    for i in range(1, 31 + 1):
        try:
            day = datetime(month.year, month.month, i)
        except ValueError:
            break

        inv_kWh_day_data = get_inverter_kwh_for_day(day, inverter)
        days.update({day.strftime("%d/%m/%Y"): inv_kWh_day_data})

        # sets first and last valid readings for month
        kWh_day_0 = inv_kWh_day_data.get("kWh_0", 0)
        if kWh_month_0 == 0 and kWh_day_0 != 0:
            kWh_month_0 = kWh_day_0
        kWh_day_1 = inv_kWh_day_data.get("kWh_1", 0)
        if kWh_day_1 != 0:
            kWh_month_1 = kWh_day_1

    kWh_total += kWh_month_1 - kWh_month_0
    inv_kWh_month_data.update({"total": kWh_total, "days": days})

    return inv_kWh_month_data


def get_inverter_kwh_for_day(day: datetime, inverter: Inverter) -> dict:
    inv_kWh_day_data = {"total": 0, "kWh_0": 0, "kWh_1": 0}

    uc = inverter.name.split("_")[0].lower()
    dir = (
        "data/{}/".format(uc)
        + str(day.year)
        + "/"
        + str(day.month)
        + "/"
        + str(day.day)
        + "/"
        + inverter.name
        + "/"
    )
    xlsx_name = day.strftime("%d_%m_%Y-") + inverter.name + ".xlsx"
    xlsx_path = Path(dir + xlsx_name)
    if not xlsx_path.exists():
        return inv_kWh_day_data

    book = openpyxl.load_workbook(xlsx_path)
    sheet = book.active

    kWh_column = 0
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == "Total Generation(kWh)":
            kWh_column = col
            break
    kWh_0 = sheet.cell(row=2, column=kWh_column).value
    kWh_1 = sheet.cell(row=sheet.max_row, column=kWh_column).value
    total = kWh_1 - kWh_0
    inv_kWh_day_data.update({"total": total, "kWh_0": kWh_0, "kWh_1": kWh_1})

    return inv_kWh_day_data


def validate_inverter_kwh_for_day(day: datetime, inverter: Inverter) -> None:
    # get kWh from server
    inverter_info = INVERTERS_INFO.get(inverter, {})
    sn = inverter_info.get("sn", 0)
    if sn == 0:
        print("Error getting inverter's sn: " + inverter.name)
        return

    url = API_URL + "v1/PowerStation/GetInverterYieldRatioChartsBySn"
    day_request = day + timedelta(days=1)
    payload = {"sn": sn, "date": str(day_request), "type": 2}
    response = requests.post(url, headers=headers, json=payload)
    data = response.json().get("data")
    pv = data.get("pv", [])
    kWh_server = 0
    for i in pv:
        date = i.get("x", "")
        if date == day.strftime("%m/%d/%Y"):
            kWh_server = i.get("y", 0)
            break

    # get kWh from local
    inv_kWh_day_data = get_inverter_kwh_for_day(day, inverter)
    kWh_local = inv_kWh_day_data.get("total", 0)
    kWh_local = float("{:.2f}".format(kWh_local))

    if kWh_server != kWh_local:
        print(
            "{} {} server:{} <> local:{}".format(
                day.date(), inverter.name, kWh_server, kWh_local
            )
        )


def get_inverter_fault_codes(inverter: Inverter, warning_msgs: dict) -> dict:
    code_dict = {}
    inverter_info = INVERTERS_INFO.get(inverter, {})
    station_id = inverter_info.get("station_id", 0)
    sn_local = inverter_info.get("sn", 0)
    if station_id == 0 or sn_local == 0:
        print("Error getting inverter's local data: " + inverter.name)
        return code_dict

    url = API_URL + "warning/PowerstationWarningsQuery"
    payload = {"pw_id": station_id}
    response = requests.post(url, headers=headers, json=payload)

    data = response.json().get("data", {})
    inverters = data.get("list", [])

    for inverter_server in inverters:
        sn_server = inverter_server.get("sn")
        if sn_server == sn_local:
            warnings = inverter_server.get("warning", [])
            codes = []
            for warning in warnings:
                warning_code = warning.get("warning_code")
                if warning_code:
                    happen_time = warning.get("happen_time")
                    codes.append(happen_time)
                    codes.append(warning_code)
                    warning_msg = warning_msgs.get(warning_code + "_warning")
                    codes.append(warning_msg)
                    reason_msg = warning_msgs.get(warning_code + "_reason")
                    codes.append(reason_msg)
                error_code = warning.get("error_code")
                if error_code:
                    codes.append(error_code)
            code_dict.update({"inverter_name": inverter.name, "fault_codes": codes})
            return code_dict

    return code_dict


def get_uc_inverter_fault_codes(uc: UC, skip: tuple) -> list:
    warning_msgs_file = open("sems_warnings.json")
    warning_msgs = json.load(warning_msgs_file)
    fault_codes = []

    for inverter in Inverter:
        if inverter in skip:
            continue

        if uc.name != inverter.name.split("_")[0]:
            continue

        inv_code_dict = get_inverter_fault_codes(inverter, warning_msgs)
        codes = inv_code_dict.get("fault_codes", [])
        if codes:
            fault_codes.append(inv_code_dict)

    return fault_codes


def main() -> None:
    skip = (Inverter.S1_BL13_2, Inverter.S1_BL4)
    login()

    # FAODO
    inverter = Inverter.FAODO_BL14
    uc = UC.FAODO
    month = datetime(2023, 12, 1)
    for i in range(1, 31 + 1):
        day = datetime(month.year, month.month, i)
        publish_uc_data_for_day(uc, day, skip)
        print("validating for day            {}".format(i))
        validate_inverter_kwh_for_day(day, inverter)
        print("\n")
    update_uc_readme_for_month(month, get_uc_kwh_for_month(uc, month, skip))

    # Get warnings
    uc_fault_codes = get_uc_inverter_fault_codes(UC.S1, skip)
    print(json.dumps(uc_fault_codes, indent=4))

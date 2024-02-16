from pathlib import Path
from datetime import datetime

import matplotlib.pyplot as plt
import openpyxl
import matplotlib.dates as mdates


def plot_inverter_kwh_for_day(xlsx_path: str) -> Path:
    book = openpyxl.load_workbook(xlsx_path)
    sheet = book.active

    # get column with power information from excel
    power_column = 0
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == "Power(W)":
            power_column = col
            break

    # get column with time information from excel
    time_column = 0
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == "Time":
            time_column = col
            break

    # axis for plot
    x = []
    y = []
    for row in range(2, sheet.max_row + 1):
        x.append(sheet.cell(row=row, column=time_column).value)
        y.append(sheet.cell(row=row, column=power_column).value / 1000)

    # customize plot
    plt.xlabel("Horário (h)")
    plt.ylabel("Potência (kW)")
    title = "Curva de geração " + str(x[0].date())
    plt.title(title)
    plt.plot(x, y)
    plt.grid()
    fmt = mdates.DateFormatter("%H")
    ax = plt.gca()
    ax.xaxis.set_major_formatter(fmt)

    # annotate max power
    ymax = max(y)
    xpos = y.index(ymax)
    xmax = x[xpos]
    annotate_text = "max=" + str(ymax) + "kW"
    ax.annotate(
        annotate_text,
        xy=(xmax, ymax),
        xytext=(xmax, ymax + 2),
        arrowprops=dict(arrowstyle="->", linewidth=1),
    )
    ax.set_ylim(0, ymax + 5)

    image_path = xlsx_path.split(".")[0] + ".png"
    fig = plt.gcf()
    fig.set_size_inches(20, 10)
    plt.savefig(image_path, bbox_inches="tight", dpi=100)
    plt.clf()
    return Path(image_path)


def plot_inverter_kwh_for_month(image_path: Path, inverter_kwh_info: dict) -> None:
    data = inverter_kwh_info.get("days")
    if data is None:
        print("data for plotting inv kwh for month is invalid")
        return
    x = []
    y = []
    for key in data:
        day = datetime.strptime(key, "%d/%m/%Y")
        x.append(day)
        y.append(data.get(key).get("total"))

    # customize plot
    plt.xlabel("Dia")
    plt.ylabel("Geração (kWh)")
    title = "Geração mensal " + datetime.strftime(x[0], "%m/%Y")
    plt.title(title)
    plt.bar(x, y)
    plt.grid(axis="y")
    fmt = mdates.DateFormatter("%d")
    ax = plt.gca()
    ax.xaxis.set_major_formatter(fmt)
    ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
    fig = plt.gcf()
    fig.set_size_inches(20, 10)
    plt.savefig(image_path, bbox_inches="tight", dpi=100)
    plt.clf()
